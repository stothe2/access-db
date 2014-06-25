from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl import cell
from openpyxl.styles import Style, Font, Alignment, Border
import datetime

class Workbook:
	'''excel documents processing'''

	def __init__(self):
		self.wbName = ''
		self.wsName = ''
		self.wsNewName = ''

	# return workbook name
	def workbook_name(self):
		return self.wbName

	# return new worksheet name 
	def worksheet_name(self):
		return self.wsNewName

	# generate error message
	def error(self, msg):
		print 'Error: ' + msg
		exit()

	# open excel file
	def load(self, wbName, wsName, wsNewName):
		if not wbName:
			self.error('invalid workbook name!')
		if not wsName:
			self.error('invalid previous worksheet name!')
		if not wsNewName:
			self.error('invalid new worksheet name!')
		if wsName == wsNewName:
			self.error('previous worksheet and new worksheet cannot be same!')
		self.wbName = wbName
		self.wsName = wsName
		self.wsNewName = wsNewName

		wb = load_workbook(filename = r'%s' % wbName)
		wsOld = wb[wsName]
		ws = wb.create_sheet()
		ws.title = wsNewName

		self.setup(wb, ws, wsOld)
		return (ws, wb)

	# edit excel file
	def setup(self, wb, ws, wsOld):
		s1 = Style(font=Font(bold=True), alignment=Alignment(horizontal='center', vertical='center'))
		s2 = Style(alignment=Alignment(horizontal='center', vertical='center'))
		s3 = Style(font=Font(bold=True))
		s4 = Style(font=Font(bold=True), alignment=Alignment(horizontal='center', vertical='center'))

		# "Med/low" table, new column
		ws['E6'].style = s1
		ws['E6'].data_type = cell.Cell.TYPE_FORMULA
		ws['E6'] = str(datetime.date.today())
		ws['E7'] = 0
		ws['E8'] = 0
		ws['E9'] = 0
		ws['E10'] = 0
		ws['E11'].data_type = cell.Cell.TYPE_FORMULA
		ws['E11'] = '=sum(E7+E8+E9+E10)'
		ws['E12'] = 0
		ws['E13'] = 0
		ws['E14'] = 0
		ws['E15'] = 0
		ws['E16'].data_type = cell.Cell.TYPE_FORMULA
		ws['E16'] = '=sum(E12+E13+E14+E15)'
		ws['E17'] = 0
		ws['E18'] = 0
		ws['E19'] = 0
		ws['E20'] = 0
		ws['E21'].data_type = cell.Cell.TYPE_FORMULA
		ws['E21'] = '=sum(E17+E18+E19+E20)'
		ws['E22'].style = s3
		ws['E22'].data_type = cell.Cell.TYPE_FORMULA
		ws['E22'] = '=sum(E11+E16+E21)'

		# "Med/low" table, old column
		ws['D6'].style = s1
		ws['D6'] = wsOld['E6'].value
		ws['D7'] = wsOld['E7'].value
		ws['D8'] = wsOld['E8'].value
		ws['D9'] = wsOld['E9'].value
		ws['D10'] = wsOld['E10'].value
		ws['D11'].data_type = cell.Cell.TYPE_FORMULA
		ws['D11'] = '=sum(D7+D8+D9+D10)'
		ws['D12'] = wsOld['E12'].value
		ws['D13'] = wsOld['E13'].value
		ws['D14'] = wsOld['E14'].value
		ws['D15'] = wsOld['E15'].value
		ws['D16'].data_type = cell.Cell.TYPE_FORMULA
		ws['D16'] = '=sum(D12+D13+D14+D15)'
		ws['D17'] = wsOld['E17'].value
		ws['D18'] = wsOld['E18'].value
		ws['D19'] = wsOld['E19'].value
		ws['D20'] = wsOld['E20'].value
		ws['D21'].data_type = cell.Cell.TYPE_FORMULA
		ws['D21'] = '=sum(D17+D18+D19+D20)'
		ws['D22'].style = s3
		ws['D22'].data_type = cell.Cell.TYPE_FORMULA
		ws['D22'] = '=sum(D11+D16+D21)'

		# "Med/low" table, formatting
		ws.merge_cells('B5:E5')
		ws.merge_cells('B6:C6')
		ws.merge_cells('B7:B11')
		ws.merge_cells('B12:B16')
		ws.merge_cells('B17:B21')
		ws.merge_cells('B22:C22')

		# "Med/low" table, text
		ws['B5'].style = s1
		ws['B5'] = 'Med/Low'
		ws['B7'].style = s2
		ws['B7'] = 'Confirmed'
		ws['B12'].style = s2
		ws['B12'] = 'Deferred'
		ws['B17'].style = s2
		ws['B17'] = 'In Review'
		ws['C7'] = 'Future'
		ws['C8'] = 'Past PCD'
		ws['C9'] = 'No PCD-Late'
		ws['C10'] = 'No PCD-OK'
		ws['C11'] = 'Subtotal'
		ws['C12'] = 'Future'
		ws['C13'] = 'Past PCD'
		ws['C14'] = 'No PCD-Late'
		ws['C15'] = 'No PCD-OK'
		ws['C16'] = 'Subtotal'
		ws['C17'] = 'Future'
		ws['C18'] = 'Past PCD'
		ws['C19'] = 'No PCD-Late'
		ws['C20'] = 'No PCD-OK'
		ws['C21'] = 'Subtotal'
		ws['B22'].style = s1
		ws['B22'] = 'Total'

		# "High" table, new column
		ws['L6'].style = s1
		ws['L6'].data_type = cell.Cell.TYPE_FORMULA
		ws['L6'] = str(datetime.date.today())
		ws['L7'] = 0
		ws['L8'] = 0
		ws['L9'] = 0
		ws['L10'] = 0
		ws['L11'].data_type = cell.Cell.TYPE_FORMULA
		ws['L11'] = '=sum(L7+L8+L9+L10)'
		ws['L12'] = 0
		ws['L13'] = 0
		ws['L14'] = 0
		ws['L15'] = 0
		ws['L16'].data_type = cell.Cell.TYPE_FORMULA
		ws['L16'] = '=sum(L12+L13+L14+L15)'
		ws['L17'] = 0
		ws['L18'] = 0
		ws['L19'] = 0
		ws['L20'] = 0
		ws['L21'].data_type = cell.Cell.TYPE_FORMULA
		ws['L21'] = '=sum(L17+L18+L19+L20)'
		ws['L22'].style = s3
		ws['L22'].data_type = cell.Cell.TYPE_FORMULA
		ws['L22'] = '=sum(L11+L16+L21)'

		# "High" table, old column
		ws['K6'].style = s1
		ws['K6'] = wsOld['L6'].value
		ws['K7'] = wsOld['L7'].value
		ws['K8'] = wsOld['L8'].value
		ws['K9'] = wsOld['L9'].value
		ws['K10'] = wsOld['L10'].value
		ws['K11'].data_type = cell.Cell.TYPE_FORMULA
		ws['K11'] = '=sum(K7+K8+K9+K10)'
		ws['K12'] = wsOld['L12'].value
		ws['K13'] = wsOld['L13'].value
		ws['K14'] = wsOld['L14'].value
		ws['K15'] = wsOld['L15'].value
		ws['K16'].data_type = cell.Cell.TYPE_FORMULA
		ws['K16'] = '=sum(K12+K13+K14+K15)'
		ws['K17'] = wsOld['L17'].value
		ws['K18'] = wsOld['L18'].value
		ws['K19'] = wsOld['L19'].value
		ws['K20'] = wsOld['L20'].value
		ws['K21'].data_type = cell.Cell.TYPE_FORMULA
		ws['K21'] = '=sum(K17+K18+K19+K20)'
		ws['K22'].style = s3
		ws['K22'].data_type = cell.Cell.TYPE_FORMULA
		ws['K22'] = '=sum(K11+K16+K21)'

		# "High" table, formatting
		ws.merge_cells('I5:L5')
		ws.merge_cells('I6:J6')
		ws.merge_cells('I7:I11')
		ws.merge_cells('I12:I16')
		ws.merge_cells('I17:I21')
		ws.merge_cells('I22:J22')

		# "High" table, text
		ws['I5'].style = s1
		ws['I5'] = 'High'
		ws['I7'].style = s2
		ws['I7'] = 'Confirmed'
		ws['I12'].style = s2
		ws['I12'] = 'Deferred'
		ws['I17'].style = s2
		ws['I17'] = 'In Review'
		ws['J7'] = 'Future'
		ws['J8'] = 'Past PCD'
		ws['J9'] = 'No PCD-Late'
		ws['J10'] = 'No PCD-OK'
		ws['J11'] = 'Subtotal'
		ws['J12'] = 'Future'
		ws['J13'] = 'Past PCD'
		ws['J14'] = 'No PCD-Late'
		ws['J15'] = 'No PCD-OK'
		ws['J16'] = 'Subtotal'
		ws['J17'] = 'Future'
		ws['J18'] = 'Past PCD'
		ws['J19'] = 'No PCD-Late'
		ws['J20'] = 'No PCD-OK'
		ws['J21'] = 'Subtotal'
		ws['I22'].style = s1
		ws['I22'] = 'Total'

		# "Grand Total" table
		ws.merge_cells('F24:H24')
		ws['F24'].style = s1
		ws['F24'] = 'Grand Total'
		ws['F26'].style = s4
		ws['F26'].data_type = cell.Cell.TYPE_FORMULA
		ws['F26'] = '=sum(D22+K22+O13+O17)'
		ws['G26'].style = s4
		ws['G26'].data_type = cell.Cell.TYPE_FORMULA
		ws['G26'] = '=sum(E22+L22+P13+P17)'
		ws['H25'].style = s1
		ws['H25'] = '%  Change'
		ws['H26'].style = s4
		ws['H26'].data_type = cell.Cell.TYPE_FORMULA
		ws['H26'] = '=(F26-G26)/F26'
		ws['F25'].style = s1
		ws['F25'] = wsOld['F25'].value
		ws['G25'].style = s1
		ws['G25'].data_type = cell.Cell.TYPE_FORMULA
		ws['G25'] = str(datetime.date.today())

		# "Linedown" table
		ws['O11'].style = s1
		ws['O11'] = 'Linedown'
		ws['O12'].style = s1
		ws['O12'] = wsOld['O12'].value
		ws['P12'].style = s1
		ws['P12'].data_type = cell.Cell.TYPE_FORMULA
		ws['P12'] = str(datetime.date.today())
		ws['O13'] = wsOld['O13'].value
		ws['P13'] = 0

		# "Safety" table
		ws['O15'].style = s1
		ws['O15'] = 'Safety'
		ws['O16'].style = s1
		ws['O16'] = wsOld['O16'].value
		ws['P16'].style = s1
		ws['P16'].data_type = cell.Cell.TYPE_FORMULA
		ws['P16'] = str(datetime.date.today())
		ws['O17'] = wsOld['O17'].value
		ws['P17'] = 0

	# save excel file
	def close(self, wb):
		wb.save(self.wbName)
